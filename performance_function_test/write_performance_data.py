import os
import shutil
import subprocess
import time
from openpyxl import Workbook,load_workbook

import xlwt
import yaml
from openpyxl import load_workbook

import sys
sys.path.append("..")
from file_util import *

si = subprocess.STARTUPINFO()
si.dwFlags |= subprocess.STARTF_USESHOWWINDOW

dev_list =[]


def creat_excel(yaml_path):
    import time
    import xlwt
    global new_excel_path
    model_excel_path = os.path.abspath( get_base_path() + 'performance_function_test\手机端性能测试报告__模板.xlsx')

    time_report = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
    excel_result = shutil.copy(model_excel_path,
                               get_base_path() + 'performance_function_test\\report\手机端性能测试结果__' + str(
                                   time_report) + '.xlsx')

    new_excel_path = os.path.abspath(excel_result)
    # 删除文件
    if os.path.exists(new_excel_path):
        os.unlink(new_excel_path)
    # 复制并重命名新文件
    shutil.copy(model_excel_path, new_excel_path)

    print('往excel中写入数据')
    workbook = load_workbook(new_excel_path)
    workbook.guess_types = True  # 猜测格式类型

    # work_book = xlwt.Workbook(encoding='utf-8')
    with open(yaml_path, 'r', encoding='utf-8') as file:
        data = yaml.load(file)
    scene_info = data['scene']
    process_info = data['process']
    lens_process = len(process_info)
    lens_scene = len(scene_info)
    print('数量是：' + str(lens_scene))

    sheet0 = workbook.get_sheet_by_name('APP性能测试')
    sheet0.cell(1, 1, '测试场景')
    n = 2
    for temp_scene in scene_info.values():
         # sheet.cell(i + begin_row, j + begin_col, float(list[j]))
         sheet0.cell(n, 1, temp_scene)
         n = n+1
    i=2
    for temp_process in process_info.values():
            sheet0.cell(1,i, '内存平均值【'+temp_process+'】')
            i=i+1
    sheet0.cell(1, lens_process + 2,'内存平均值【所有加起来】')
    temp_lens_process =lens_process+3
    for temp_process in process_info.values():
        sheet0.cell(1, temp_lens_process,'CPU平均值【'+temp_process+'】')
        temp_lens_process=temp_lens_process+1

    sheet0.cell(1 , lens_process*2 + 3, '内存CPU【所有加起来】')


    for i in range(1, 6):
        sheeti = workbook.get_sheet_by_name('第' + str(i) + '次测试')
        # print(sheeti)
        sheeti.cell(1, 1, '测试场景')
        sheeti.cell(2, 1, '指标')
        sheeti.cell(3, 1, '平均值')

        n = 2
        for temp_scene in scene_info.values():
            sheeti.cell(1,n,temp_scene)

            # 插入场景进程
            j = 1+(n-1)
            for temp_process in process_info.values():
                sheeti.cell(2,j, temp_process + '的内存是：（M）')
                j = j + 1
            sheeti.cell(2,j, '总的内存是：（M）')

            j = j + 1
            for temp_process in process_info.values():
                 sheeti.cell(2,j, temp_process + '的CPU是：（%）')
                 j = j + 1
            sheeti.cell(2,j,'总的CPU是：（M）')

            n = n + (int(lens_process) + 1) * 2
    workbook.save(new_excel_path)
    workbook.close()
    return new_excel_path

def write_data_to_excel(new_txt_path, excel_path, begin_row, begin_col, sheet_name):
    '''
    :param old_txt_path:  新的的txt数据值路径
    :param excel_path: excel文件绝对路径
    :param begin_row: 写入数据开始行号，最小为1
    :param begin_col: 写入数据开始列数，最小为1
    :return:
    '''
    file_txt = open(new_txt_path)
    lines = file_txt.readlines()

    if (begin_row <= 1):
        begin_row = 1
    if (begin_col <= 1):
        begin_col = 1
    workbook = load_workbook(excel_path)
    workbook.guess_types = True  # 猜测格式类型
    sheet = workbook.get_sheet_by_name(sheet_name)

    if (len(lines) == 0):
        print("没有数据.......")
        return

    total_value = None
    for i in range(len(lines)-1):
        list = (lines[i].split(','))
        if (len(list) == 0):
            continue

        if total_value==None:
            total_value = [0] * len(list)

        for j in range(len(list)):
            # print('temp的数据 ' + list[j])
            sheet.cell(i + begin_row, j + begin_col, float(list[j]))
            total_value[j] = total_value[j] + float(list[j])

    average_value = []
    if total_value != None:
        average_value = [0] * len(total_value)
        for k in range(len(total_value)):
            average_value[k] =  float(total_value[k])/(len(lines)-1)
            if begin_row<=1:
                break
            sheet.cell(begin_row -1 , k + begin_col, average_value[k])

    workbook.save(excel_path)
    workbook.close()
    file_txt.close()
    # 返回平均值列表
    return average_value


def write_excel(begin_row, begin_col, sheet_name):
    txt_path = os.path.abspath(get_base_path() + 'performance_function_test\log\log.txt')
    excel_path = creat_excel()
    write_data_to_excel(txt_path,excel_path,begin_row, begin_col, sheet_name)
    time.sleep(5)
    print('写入成功')


if __name__ == "__main__":
    # creat_excel()
    write_excel(4,2,'第1次测试')
