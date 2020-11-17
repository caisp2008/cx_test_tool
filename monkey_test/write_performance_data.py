import os
import shutil
import yaml
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference

import sys
sys.path.append("..")
from file_util import *

def write_data_to_excel(txt_path,excel_path,begin_row, begin_col):
    '''
    :param old_txt_path:  旧的txt数据值路径
    :param excel_path: excel文件绝对路径
    :param begin_row: 写入数据开始行号，最小为1
    :param begin_col: 写入数据开始列数，最小为1
    :return:
    '''
    file_txt = open(txt_path)
    lines = file_txt.readlines()

    if (begin_row <= 1):
        begin_row = 1
    if (begin_col <= 1):
        begin_col = 1

    workbook = load_workbook(excel_path)
    workbook.guess_types = True  # 猜测格式类型
    sheet = workbook[u"测试数据采集"]


    if (len(lines) == 0):
        print("没有数据.......")
        return


    for i in range(len(lines)):
        list = (lines[i].split(','))
        if (len(list) == 0):
            continue

        for j in range(len(list)):
            # print('temp的数据 ' + list[j])
            sheet.cell(i + begin_row, j  + begin_col, float(list[j]))
    workbook.save(excel_path)
    workbook.close()
    file_txt.close()
    return lines,sheet

def creat_excel():
    import time
    model_excel_path = os.path.abspath(get_base_path() + 'monkey_test\\手机端monkey性能测试结果_模板.xlsx')
    time = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
    kugou_result = shutil.copy(model_excel_path, get_base_path() + 'Folder\APP_monkey_folder\\report\\手机端monkey性能测试结果_' + str(time) + '.xlsx')

    new_excel_path=os.path.abspath(kugou_result)
    print(new_excel_path)
    # 删除文件
    if os.path.exists(new_excel_path):
        os.unlink(new_excel_path)
    # 复制并重命名新文件
    shutil.copy(model_excel_path, new_excel_path)
    return new_excel_path

def write_to_excel(txt_path, excel_path, begin_row, begin_col):
    '''
    :param txt_path:  txt文件绝对路径
    :param excel_path: excel文件绝对路径
    :param begin_row: 写入数据开始行号，最小为1
    :param begin_col: 写入数据开始列数，最小为1
    :return:
    '''
    file_txt=open(txt_path)

    workbook = load_workbook(excel_path)
    workbook.guess_types = True  # 猜测格式类型
    chart_sheet =workbook[u"测试图形结果"]
    data_sheet = workbook[u"测试数据采集"]
    lines = file_txt.readlines()

    if (begin_row <= 1):
        begin_row = 1
    if (begin_col <= 1):
        begin_col = 1

        # 用with的写法是就算没有关闭文件夹，也可以打开文件
    with open(get_base_path() + 'monkey_test\\yaml\\performance_caps.yaml', 'r', encoding='utf-8') as file:
        data = yaml.load(file)
    print(data)
    process_info = data['process']



    for i in range(1, 4):
        sheeti = workbook.get_sheet_by_name('测试数据采集')
        j = 2
        for temp_process in process_info.values():
            sheeti.cell(1, j, temp_process + '的内存是：（M）')
            j = j + 1
        sheeti.cell(1, j, '总的内存是：（M）')

        h = 2+len(process_info)+1
        for temp_process in process_info.values():
            sheeti.cell(1, h, temp_process + '的CPU是：（%）')
            h = h + 1
        sheeti.cell(1, h, '总的CPU是：（M）')

    def get_column_name(index):
        '''
        获取Excel列名，例如 A、B、AB
        :param index: 大于0的数字
        :return:
        '''

        # 向下取整
        index_1 = (index - 1) // 26
        # 求余数
        index_2 = (index - 1) % 26

        if index_1 == 0:
            return chr(ord('A') + index - 1)
        else:
            return chr(ord('A') + index_1) + chr(ord('A') + index_2)

    # def write_new_data(data_sheet, chart_sheet, data_min_row, data_min_col, data_max_row, data_max_col, title,chart_point):
    n=0
    j=0

    for temp_process in process_info.values():
        write_new_data(data_sheet,chart_sheet, begin_row, begin_col+n, begin_row+len(lines),begin_col+n,temp_process+"内存（M）",get_column_name(11+10*j) + str(21*0+3))
        write_new_data(data_sheet,chart_sheet, begin_row, begin_col+len(process_info)+1+n, begin_row+len(lines), begin_col+len(process_info)+1+n ,temp_process+"的CPU(%)",get_column_name(11+10*j) + str(21*1+5))
        n=n+1
        j=j+1
    write_new_data(data_sheet, chart_sheet, begin_row, begin_col + len(process_info), begin_row + len(lines),
                   begin_col + len(process_info) , "总的内存（M）", "A" + str(21 * 0 + 3))
    write_new_data(data_sheet, chart_sheet, begin_row, begin_col+len(process_info)*2+1, begin_row + len(lines),begin_col+len(process_info)*2+1, "总的CPU (%)", "A" + str(21 * 1 + 5))
    workbook.save(excel_path)
    workbook.close()
    file_txt.close()


def write_new_data(data_sheet, chart_sheet, data_min_row, data_min_col, data_max_row, data_max_col, title, chart_point):
    c1 = LineChart()
    c1.title = title  # 图的标题
    c1.style = 11  # 线条的style
    # c1.y_axis.title = 'percent'  # y坐标的标题
    if 'IDC' not in title:
        c1.x_axis.number_format = 'd-mmm'  # 规定日期格式  这是月,年格式
        c1.x_axis.majorTimeUnit = "Months"  # 规定日期间隔 注意days；Months大写
    # c1.x_axis.title = "Date"  # x坐标的标题

    data = Reference(data_sheet, min_row=data_min_row, min_col=data_min_col, max_row=data_max_row, max_col=data_max_col)  # 图像的数据 起始行、起始列、终止行、终止列
    c1.add_data(data, titles_from_data=False, from_rows=False)
    dates = Reference(chart_sheet, min_col=10, min_row=1, max_col=14,max_row=20)
    c1.set_categories(dates)
    chart_sheet.add_chart(c1, chart_point)  # 将图表添加到 sheet中

def write_kugouring_excel():
    txt_path = os.path.abspath(get_base_path() + 'monkey_test\\log\\adb_log.txt')

    excel_path = creat_excel()
    write_data_to_excel(txt_path, excel_path, 5, 2)

    write_to_excel(txt_path, excel_path, 5, 2)
    print('写入成功')

if __name__ == '__main__':

    write_kugouring_excel()

