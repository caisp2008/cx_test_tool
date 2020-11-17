# -*- coding: utf-8 -*-
import os
import threading
import tkinter as tk # imports
from tkinter import ttk
# from performance_function_test.write_performance_data import getPerformance_data
from tkinter import messagebox
import time
import yaml
import inspect
import ctypes

from openpyxl import load_workbook

# from performance_function_test.write_performance_data import write_performance_data
from performance_function_test.write_performance_data import creat_excel, write_data_to_excel

import sys
sys.path.append("..")
from file_util import *

global projectid,path_report
projectid=1
path_report=get_base_path() + 'performance_function_test\\report'
path_yaml = get_base_path() + 'performance_function_test\\yaml'

def init_tab2(tab_page):
    global text_monkey_info
    monty = ttk.LabelFrame(tab_page, text=' ')
    monty.grid(column=0, row=0, padx=20, pady=20)

    tk.Button(monty, text='点击配置文件 ', font=("Arial", 11), width=30, height=1,command=lambda: openReport(path_yaml)).grid(column=0, row=0, sticky='W', padx=20, pady=10)
    tk.Button(monty, text='点击开始运行 ', font=("Arial", 11), width=30, height=1,command=run_function_test).grid(column=0, row=5, sticky='W', padx=20,
                                                                                   pady=10)
    tk.Button(monty, text='查看测试结果 ', font=("Arial", 11), width=30, height=1,command=lambda: openReport(path_report)).grid(column=0, row=9, sticky='W', padx=20,
                                                                                   pady=10)

def get_function():
    log_path = get_base_path() + 'performance_function_test\\log\\log.txt'
    path_log = os.path.abspath(log_path)
    excel_path = creat_excel(path_yaml)


    with open(get_base_path() + 'performance_function_test\\yaml\\performance_caps.yaml', 'r', encoding='utf-8') as file:
        data = yaml.load(file)
    scene_info = data['scene']
    # print(scene_info)
    process_info = data['process']
    lens_process = len(process_info)
    lens_scene = len(scene_info)

    averageValueLog_path = get_base_path() + 'performance_function_test\log\\averageValueLog.txt'
    path_averageValueLog = os.path.abspath(averageValueLog_path)
    file_averageValueLog= open(path_averageValueLog, 'w')

    isCustomStop = False # 是否用户取消
    all_times_total = [0] * len(scene_info) # 保存每次的平均值总和，每个item是一个场景的平均值列表
    has_value_times = [0] * len(scene_info) # 保存每个场景有数据的次数是多少次
    all_times_average = [0] * len(scene_info)
    for n in range(1,4):
        # ask = messagebox.askquestion('APP性能测试', '是否开始第 '+ str(n)+' 轮测试（测试会有3轮测试）')
        # print(ask)
        # while askquestion == 'yes':

        y =2
        scene_index =0
        for temp_scene in scene_info.values():
            # f = open(path_log, 'w')
            # f.closed
            f = open(path_log, 'w')

            start_info = tk.messagebox.askquestion(title='第'+ str(n) +'轮测试：' + temp_scene,message='请对页面进行操作，操作完成后，点击“确定”，进行数据采集')
            print(start_info)
            th = threading.Thread(target=set_text_test, args=(f,))
            if start_info=='yes':
                print('数据采集中')
                th.setDaemon(True)  # 守护线程
                th.start()
            else:
                isCustomStop = True
                break
            end_info=tk.messagebox.showinfo('数据采集中', '正在采集数据，请至少采集30S的数据，\n点击" 确定 "可结束数据采集')

            # tk.Label(monty,text='路径为：', font=("Arial", 11), width=20, height=2).grid(column=1,row=6, sticky='W')
            if end_info =='ok':
                _async_raise(th.ident,SystemExit)
                print('关闭txt文件')
                f.close()
                averageValue=write_data_to_excel(path_log, excel_path, 4, y, '第' + str(n) + '次测试')

                if len(averageValue)>0:
                    has_value_times[scene_index] = has_value_times[scene_index] + 1

                tmp = all_times_total[scene_index]
                if type(tmp).__name__ == 'list' and len(tmp)==len(averageValue):
                    for m in range(len(tmp)):
                        tmp[m] =tmp[m] + averageValue[m]
                    all_times_total[scene_index] = tmp
                else:
                    all_times_total[scene_index]=averageValue


            print('写入成功')
            scene_index = scene_index +1
            y = y + lens_process*2 + 2

            with open('结果存放.txt', 'a') as file_handle:  # .txt可以不自己新建,代码会自动新建
                file_handle.write('\n')  # 有时放在循环里面需要自动转行，不然会覆盖上一条数据
        # print(file=file_averageValueLog, end='\n')
        if isCustomStop:
            break
    if not isCustomStop:
        tk.messagebox.showinfo('APP性能测试', '测试已完成，请查看测试结果')

        # 求总平均值
        for v1 in range(len(all_times_total)):
            tmp = [0] * len(all_times_total[v1])

            for v2 in range(len(all_times_total[v1])):
                if has_value_times[v1] == 0:
                    break
                tmp[v2] = all_times_total[v1][v2]/has_value_times[v1]
            all_times_average[v1] =tmp
        print("has_value_times:",has_value_times)
        print("all_times_average:",all_times_average)

        print('将平均值写入excel')

        workbook = load_workbook(excel_path)
        workbook.guess_types = True  # 猜测格式类型
        sheet_average = workbook.get_sheet_by_name('APP性能测试')

        begin_row = 2
        begin_col = 2

        for i in range(len(all_times_average)):
            if (len(all_times_average[i]) == 0):
                continue
            for j in range(len(all_times_average[i])):
                sheet_average.cell(row=begin_row + j, column=begin_col + i, value=round(all_times_average[i][j], 2))

        workbook.save(excel_path)
        workbook.close()

    # file_averageValueLog.close()
# #运行多线程
# def get_function(project_id,insert_text):
#     getPerformance_data(project_id,insert_text)

def set_text_test(f):
    global get_adb_kugou_isrun
    get_adb_kugou_isrun = True

    while get_adb_kugou_isrun:
        get_adb_info(f)




def get_adb_kugou_stop(log_path,excel_path,n,y):
    global get_adb_kugou_isrun
    get_adb_kugou_isrun = False
    write_data_to_excel(log_path, excel_path, 4, y, '第' + str(n) + '次测试')
    # insert_text( '第 ' + str(n) + ' 次测试__' + '场景 ' + str(y) + ' 测试结束了\n')



def run_function_test():
    th = threading.Thread(target=get_function())
    th.setDaemon(True)  # 守护线程
    th.start()

#往txt中插入数据
def insert_text(message):
    if(message==None):
        message = ""

    text_monkey_info.insert(tk.END,message)

def openReport(path_report):

    path = os.path.abspath(path_report)  # 获取当前脚本所在的路径
    os.startfile(path)

def _async_raise(tid, exctype):
   """raises the exception, performs cleanup if needed"""
   tid = ctypes.c_long(tid)
   if not inspect.isclass(exctype):
      exctype = type(exctype)
   res = ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, ctypes.py_object(exctype))
   if res == 0:
      raise ValueError("invalid thread id")
   elif res != 1:
      # """if it returns a number greater than one, you're in trouble,
      # and you should call it again with exc=NULL to revert the effect"""
      ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, None)
      raise SystemError("PyThreadState_SetAsyncExc failed")

