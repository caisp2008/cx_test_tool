# -*- coding: utf-8 -*-
import re
import threading
import tkinter as tk # imports
from tkinter import ttk
import os
import subprocess
import numpy as np
import xlrd
import yaml
import time
import datetime
from numpy import *
from PIL import Image
import glob

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from file_util import get_base_path
from mytkinter.PageTab.PageTab2 import _async_raise

from performance_function_test.write_performance_data import creat_excel, write_data_to_excel

global projectid,path_report,path_yaml
projectid=1
path_report=get_base_path() + 'performance_function_test\\report'
path_yaml = get_base_path() + 'Folder\APP_xn_auto_folder\yaml '
path_air = get_base_path() + 'Folder\APP_xn_auto_folder\\air'
path_error_video = get_base_path() + 'mytkinter\VideoPhoto'
path_package = get_base_path() + 'Folder\APP_xn_auto_folder\package'


def init_tab5(tab_page):
    global text_monkey_info
    monty = ttk.LabelFrame(tab_page, text=' ')
    monty.grid(column=0, row=0, padx=20, pady=20)


    # tk.Button(monty, text='点击查看教程', font=("Arial", 12), width=30, height=1,command = open_course).grid(column=1, row=1, sticky='W', padx=20,
    #                                                                            pady=10)


    tk.Button(monty, text='点击配置文件', font=("Arial", 12), width=30, height=1,command=lambda: open_world(path_yaml)).grid(column=1, row=3, sticky='W', padx=20,
                                                                               pady=10)

    tk.Button(monty, text='点击将脚本文件复制到此文件夹内', font=("Arial", 12), width=30, height=1,command=lambda: open_world(path_air)).grid(column=1, row=2, sticky='W', padx=20,
                                                                                 pady=10)

    # tk.Button(monty, text='请将测试包放在该文件夹\n并删除之前的包', font=("Arial", 12), width=30, height=2,command=lambda: open_world(path_package)).grid(column=1, row=4, sticky='W', padx=20,
    #                                                                              pady=10)

    # tk.Button(monty, text='点击安装测试包', font=("Arial", 12), width=30, height=2,command=install_package).grid(column=2, row=4,
    #                                                                                            sticky='W', padx=20,
    #                                                                                            pady=10)


    tk.Button(monty, text='点击开始测试', font=("Arial", 12), width=30, height=1,command=run_start_APP_auto_test).grid(column=1, row=6, sticky='W', padx=20,
                                                                                 pady=10)
    tk.Button(monty, text='点击查看测试结果', font=("Arial", 12), width=30, height=1,command=lambda: open_world(path_report)).grid(column=2, row=6, sticky='W', padx=20,
                                                                                 pady=10)

    tk.Button(monty, text='点击查看失败截图', font=("Arial", 12), width=30, height=1,
              command=lambda: open_world(path_error_video)).grid(column=3, row=6, sticky='W', padx=20,
                                                            pady=20)

    text_monkey_info = tk.Text(monty, width=125, height=16, font=("Arial", 12))
    text_monkey_info.grid(row=8, column=0, sticky='W', columnspan=7, padx=10, pady=10)


#开始自动化测试方法
def start_APP_auto_test():
    insert_text('自动化测试开始了\n')
    tk.messagebox.showinfo('测试开始提示', '测试开始前，开启itest，请检查itest账号是否登录，权限是否给与，确定给与后点击确认开始测试')
    # cmd = "adb shell am start -a android.intent.action.MAIN -c android.intent.category.LAUNCHER –n iflytek.testTech.propertytool/.activity.BootActivity"
    # print("运行脚本是：" + cmd)
    # air_result = os.system(cmd)
    #检测手机是否连接成功
    testDevices()

    # 开启悬浮窗权限
    cmd = 'adb shell am broadcast -a enableFloatWindow'
    os.system(cmd)

    #将失败截图中的照片清除
    path = get_base_path() + 'mytkinter\VideoPhoto'
    paths = glob.glob(os.path.join(path, '*.png'))
    # 输出所有文件和文件夹
    for file in paths:
        fp = open(file, 'rb')
        img = Image.open(fp)
        fp.close()
        os.remove(file)
        print('照片删除成功')

    log_path_air_fail = get_base_path() + 'Folder\APP_xn_auto_folder\log'
    path_air_fail_log = os.path.abspath(log_path_air_fail)
    f_air_log = open(path_air_fail_log, 'w')

    #打开yaml文件，并获取包名，进程名等信息
    with open(get_base_path() + 'Folder\APP_xn_auto_folder\yaml\APP_xn_auto.yaml', 'r', encoding='utf-8') as file:
        data = yaml.load(file)
    script_info = data['script']
    process_info = data['process']
    lens_process = len(process_info)
    scene_info = data['scene']
    # 场景数量
    lens_scene = len(scene_info)

    packageName_info = data['packageName']
    no_data_script_info = data['no_need_collect_data_script']
    print(len(process_info))
    print(script_info.values())

    insert_text('创建excel表格\n')
    log_path = get_base_path()  + 'performance_function_test\\log\\log.txt'
    path_log = os.path.abspath(log_path)
    yaml_path = 'Folder\APP_xn_auto_folder\yaml\APP_xn_auto.yaml'
    excel_path = creat_excel(yaml_path)
    insert_text('创建表格成功\n')

    all_average_pss = []
    all_average_cpu = []

    #开始测试采集数据的脚本
    for n in range(1, 6):
        len_process = 0
        average_process = 0
        # 测试非数据采集场景
        testDevices()
        for temp_no_data_script in no_data_script_info.values():
            path = get_base_path() + 'Folder\APP_xn_auto_folder\\air\\' + temp_no_data_script
            print('测试的脚本路径是：' + path)
            cmd = "airtest run " + path + "  --device Android:/// --log log/"
            print("运行脚本是：" + cmd)
            air_result = os.system(cmd)
            if air_result == 0:
                print('测试通过')
            else:
                screenshot('第'+ str(n) + '轮测试' + temp_no_data_script)
                # print(temp_no_data_script+'测试出错',file = f_air_log,end='\n')
                insert_text('第'+ str(n) + '轮测试：'+temp_no_data_script + '测试出错，进行重新跑脚本\n')
                air_result
                insert_text('第' + str(n) + '轮测试：' + temp_no_data_script + '脚本重试失败，请查看失败截图\n')


            # print('测试脚本' + 'temp_no_data_script' + '测试完成了，干掉APP')
            # os.system("adb shell am force-stop  " + packageName_info)

        #开始测试需要采集性能数的场景
        testDevices()
        y = 2
        for temp_script in script_info.values():
            path = get_base_path()  + 'Folder\APP_xn_auto_folder\\air\\' + temp_script
            print('测试的脚本路径是：' + path)
            cmd = "airtest run " + path + "  --device Android:/// --log log/"
            print("运行脚本是：" + cmd)
            air_result = os.system(cmd)

            if air_result == 0:
                print('测试通过')
                test_result = 1
            else:
                screenshot('第'+ str(n) + '轮测试' + temp_no_data_script)
                insert_text('第'+str(n) + '轮测试：' + temp_script + '测试出错，进行脚本重试\n')
                # 定义测试失败的时候，test_result = 0 ，后续在失败的地方输入测试结果进行标红
                air_result
                insert_text('第' + str(n) + '轮测试：' + temp_script + '脚本重试失败，请查看失败截图\n')

                test_result = 0

            print('操作完成，现在等待30S进入稳定态')
            time.sleep(30)
            print('等待完成，现在采集数据30S')

            #开启itest进行数据采集
            #获取进程的数据，组合成itest可以使用的adb命令多进程
            process_info_all = ''
            for temp_process_info in process_info.values():
                process_info_all = temp_process_info + ',' + process_info_all
            print(process_info_all[0:-1])

            testDevices()
            # 开启itest权限
            cmd = "adb shell dalvikvm -cp /sdcard/start.dex Start"
            os.system(cmd)
            # 开始测试多进程
            cmd = "adb shell am broadcast -a monitorStart --es monitor cpu,pss --es pkg " + process_info_all[0:-1] + "  --ei interval 1000"
            os.system(cmd)

            #采集数据20S
            time.sleep(20)

            #采集数据20S之后结束数据采集
            cmd = "adb shell am broadcast -a monitorFinish"
            os.system(cmd)

            cmd = "adb shell am broadcast -a monitorFinish"
            os.system(cmd)

            #将日志信息填写到excel中
            get_handtest_log_info(excel_path, n, process_info, temp_script, lens_process, len_process, average_process, all_average_pss, all_average_cpu,test_result)

            len_process = len_process + 2 + lens_process + lens_process
            average_process = len_process


            # print('场景测试完成，干掉进程')
            # os.system("adb shell am force-stop " + packageName_info)

        print('一轮测试完成，清除缓存进行下一轮测试')
        os.system("adb shell pm clear " + packageName_info)

    #算出所有组数的平均值并写入到excel中

    # 读取excel数据
    file = xlrd.open_workbook(excel_path)

    # 写入excel
    workbook = load_workbook(excel_path)
    workbook.guess_types = True  # 猜测格式类型
    sheet_average = workbook.get_sheet_by_name('APP性能测试')

    for j in range(0, (lens_process * 2 + 2) * lens_scene):
        zushu = []
        for test_times in range(1, 6):
            sheet = file.sheet_by_name("第" + str(test_times) + "次测试")

            # 获取第一行的数据
            row_value = sheet.cell_value(2, j + 1)
            # print(row_value)
            zushu.append(row_value)
        print('第' + str(j + 1) + '组数' + str(zushu))

        print('最大值' + str(max(zushu)))
        print('最小值' + str(min(zushu)))
        pingjun = (sum(zushu) - max(zushu) - min(zushu)) / 3
        print('去除最大值和最小值后的平均值是：'+ str(pingjun))
        sheet_average.cell(2 + j // (lens_process * 2 + 2), 2 + j % (lens_process * 2 + 2), pingjun)

    workbook.save(excel_path)
    workbook.close()

    #判断测试过程中是否出现了测试失败的情况，如果有失败就打印出来，无失败就提示无失败
    path_air_log_txt = get_base_path() + 'Folder\APP_xn_auto_folder\log'
    file_txt = open(path_air_log_txt)
    lines = file_txt.readlines()

    tk.messagebox.showinfo('测试完成提示', '测试完成，请点击查看测试结果')


def write_average_to_excel(new_txt_path, excel_path, begin_row, begin_col, sheet_name):
        '''
        :param old_txt_path:  新的的txt数据值路径
        :param excel_path: excel文件绝对路径
        :param begin_row: 写入数据开始行号，最小为1
        :param begin_col: 写入数据开始列数，最小为1
        :return:
        '''
        file_txt = open(new_txt_path)
        lines = file_txt.readlines()

        if (begin_row <= 1):
            begin_row = 1
        if (begin_col <= 1):
            begin_col = 1
        workbook = load_workbook(excel_path)
        workbook.guess_types = True  # 猜测格式类型
        sheet = workbook.get_sheet_by_name(sheet_name)

        if (len(lines) == 0):
            print("没有数据.......")
            return

        total_value = None
        for i in range(len(lines)):
            list = (lines[i].split(','))
            if (len(list) == 0):
                continue

            if total_value == None:
                total_value = [0] * len(list)

            for j in range(len(list)):
                # print('temp的数据 ' + list[j])
                sheet.cell(i + begin_row, j + begin_col, float(list[j]))
                total_value[j] = total_value[j] + float(list[j])

        average_value = []
        if total_value != None:
            average_value = [0] * len(total_value)
            for k in range(len(total_value)):
                average_value[k] = float(total_value[k]) / (len(lines))
                if begin_row <= 1:
                    break
                sheet.cell(begin_row, k + begin_col, average_value[k])

        workbook.save(excel_path)
        workbook.close()
        file_txt.close()
        # 返回平均值列表
        return average_value


def run_start_APP_auto_test():
    th = threading.Thread(target=start_APP_auto_test)
    th.setDaemon(True)  # 守护线程
    th.start()


def open_world(path):

    path = os.path.abspath(path)  # 获取当前脚本所在的路径
    os.startfile(path)

# 截图函数
def screenshot(air_name):
        # print('kill ADB')
        # cmd = "adb kill-server "
        # os.system(cmd)
        # print('重启 ADB')
        # cmd = "adb start-server "
        # os.system(cmd)

        path_video = get_base_path() + 'mytkinter\VideoPhoto'
        nowtime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        name = air_name +'_'+ nowtime + '.png'
        print('name:' + name)
        replace_name = name.replace('\\', '_')
        print('replace_script---------' + replace_name)


        screen_cmd = "adb shell /system/bin/screencap -p /sdcard/screenshot.png"
        os.system(screen_cmd)

        videoPhont_path =  get_base_path() + 'mytkinter\VideoPhoto'
        pull_computer_cmd = 'adb pull /sdcard/screenshot.png ' + videoPhont_path
        os.system(pull_computer_cmd)

        rename_photo_name =videoPhont_path +'\\' + replace_name
        print('rename_photo_name:' + rename_photo_name)
        time.sleep(2)
        os.rename(videoPhont_path + '\screenshot.png', rename_photo_name)

        # out1 = subprocess.getstatusoutput('adb shell / system / bin / screencap - p / sdcard /  '+ path_video+ '/{}'.format(name))
        # print(out1)

        # if out1[0] == 0:
        #     pass
        # else:
        #     print('截图失败，失败原因：'+ out1)
        #     insert_text('未知原因，截图失败\n')


def install_package_tab():
    file_dir = get_base_path() + 'Folder\APP_xn_auto_folder\package'
    for root, dirs, files in os.walk(file_dir):
        print(files)  # 当前路径下所有非目录子文件
        print(dirs)  # 当前路径下所有子目录
        file_package = file_dir + '\\' + files[0]
        cmd = "adb install " + file_package
        print("运行脚本是：" + cmd)
        os.system(cmd)

def install_package():
    th = threading.Thread(target=install_package_tab)
    th.setDaemon(True)  # 守护线程
    th.start()

#打开使用教程
def open_course_word():
    path = get_base_path() + 'Folder\APP_xn_auto_folder\Course_word\\xn_course.docx'
    os.system(path)

#开启教程多线程
def open_course():
    th = threading.Thread(target=open_course_word)
    th.setDaemon(True)  # 守护线程
    th.start()

#分析日志信息
def get_handtest_log_info(excel_path,n,process_info,temp_script,lens_process,len_process,average_process,all_average_pss,all_average_cpu,test_result):
    # 获取存放handTest日志路径
    path_handTest_log = get_base_path() + 'Folder\performance_function_test_H_folder\log'
    print('存放handTest日志路径在这里。。。。。。。。。。。'+path_handTest_log)
    # 将手机上的handtest日志拷贝到电脑上来
    cmd_cpu = "adb pull /storage/emulated/0/AndroidPropertyTool4/handTest " + path_handTest_log
    os.system(cmd_cpu)

    # 拿出日志handtest日志文件，并将日志信息填写到excel中
    workbook = load_workbook(excel_path)
    workbook.guess_types = True  # 猜测格式类型
    sheet = workbook.get_sheet_by_name('第' + str(n) + '次测试')
    average_pss = []
    average_cpu = []

    for temp_process_info in process_info.values():
        # print(temp_process_info)
        process_list = temp_process_info.split('.')

        path_process = ''
        for temp_process in process_list:
            # print(temp_process)
            path_process = path_process + '_' + temp_process

        # 计算一次测试的数据平均值与总和
        pss_path = "pss" + path_process + ".txt"
        once_pss = []
        print("--------------------------------")

        # 获取handTest的txt文件路径
        # path_handTest_txt = path_handTest_log + '\\'
        path_handTest_txt = path_handTest_log + '\handTest\\'

        print('handTest的txt文件路径是。。。。。。。'+path_handTest_txt)
        once_pss_lines = open(path_handTest_txt + pss_path).readlines()

        for line in range(2, len(once_pss_lines)):
            data = once_pss_lines[line].split()
            if test_result == 0:
               sheet.cell(2 + line, 2 + len_process, float(data[1]))

               #运行过程中出错的话就填充红色
               orange_fill = PatternFill(fill_type='solid', fgColor="FFFF0000")
               sheet.cell(row=2 + line, column=2 + len_process).fill = orange_fill


            else:
                sheet.cell(2 + line, 2 + len_process, float(data[1]))
            once_pss.append(data[1])

        print('第' + str(n) + '轮: ' + '脚本' + temp_script + temp_process_info + '进程一组数据的内存值' + str(once_pss))
        sum = 0
        for i in once_pss:
            sum = sum + int(i)
        b = sum / int(len(once_pss))
        print('第' + str(n) + '轮: ' + '场景' + temp_script + temp_process_info + '进程一组数据的内存平均值是' + str(int(b)))
        average_pss.append(float(b))

        # 计算一次测试的CPU数据平均值与总和
        cpu_path = "cpu" + path_process + ".txt"
        once_cpu = []

        once_cpu_lines = open(
            path_handTest_txt + cpu_path).readlines()

        for temp_once_cpu_lines in range(2, len(once_cpu_lines)):
            data_cpu = once_cpu_lines[temp_once_cpu_lines].split()
            if test_result == 0:
                sheet.cell(2 + temp_once_cpu_lines, 2 + 1 + lens_process + len_process, float(data_cpu[1]))

                # 运行过程中出错的话就填充红色
                orange_fill = PatternFill(fill_type='solid', fgColor="FFFF0000")
                sheet.cell(row=2 + temp_once_cpu_lines, column=2 + 1 + lens_process + len_process).fill = orange_fill


            else:
                sheet.cell(2 + temp_once_cpu_lines, 2 + 1 + lens_process + len_process, float(data_cpu[1]))


            once_cpu.append(data_cpu[1])

        sum = 0
        print('第' + str(n) + '轮: ' + '场景' + temp_script + temp_process_info + '进程一组数据的CPU值：' + str(once_cpu))
        for i in once_cpu:
            # print(i)
            sum = sum + float(i)

        once_cpu_info = sum / int(len(once_cpu))
        print('第' + str(n) + '轮: ' + '场景' + temp_script + temp_process_info + '进程一组数据的CPU平均值是' + str(int(once_cpu_info)))
        average_cpu.append(float(once_cpu_info))
        len_process = len_process + 1

        workbook.save(excel_path)
        workbook.close()

    pss_sum_numbers = 0
    for temp_average_pss in average_pss:
        # 将一组数据内存平均值写入到excel中
        sheet.cell(3, 2 + average_process, temp_average_pss)
        pss_sum_numbers += temp_average_pss
        average_process = average_process + 1
    average_pss.append(int(pss_sum_numbers))

    sheet.cell(3, 2 + average_process, pss_sum_numbers)
    average_process = average_process + 1
    print('内存平均值列表' + str(average_pss))

    cpu_sum_numbers = 0
    for temp_average_cpu in average_cpu:
        # 将一组数据的cpu平均值值写入到excel中
        sheet.cell(3, 2 + average_process, temp_average_cpu)
        cpu_sum_numbers += temp_average_cpu
        average_process = average_process + 1

    average_cpu.append(cpu_sum_numbers)
    # 将一组CPU数据的所有平均值总和写入到excel中
    sheet.cell(3, 2 + average_process, cpu_sum_numbers)
    print('cpu的平均值列表' + str(average_cpu))

    average_process = average_process + 1

    all_average_pss.append(average_pss)
    all_average_cpu.append(average_cpu)

    print('所有组数的内存平均值列表：' + str(all_average_pss))
    print('所有组数的CPU平均值列表：' + str(all_average_cpu))
    workbook.save(excel_path)
    workbook.close()

#检测手机是否连接成功
def testDevices():
    ##获取设备多台设备号列表
    str_init = ' '
    all_info = os.popen('adb devices').readlines()
    print('adb devices 输出的内容是：', all_info)

    for i in range(len(all_info)):
        str_init += all_info[i]
    devices_name = re.findall('\n(.+?)\t', str_init, re.S)

    if len(devices_name) > 0:
        print('所有设备名称：\n', devices_name)
        return devices_name
    else:
        tk.messagebox.showerror(title='手机连接失败提示', message='手机连接失败了，请重新连接手机并点击确定继续测试')

def insert_text(message):
    now_time = datetime.datetime.now()
    text_monkey_info.insert(tk.END, '[ '+ str(now_time) + ' ] '+ message)

if __name__ == "__main__":
    insert_text('lalalla')


