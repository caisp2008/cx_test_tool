# -*- coding: utf-8 -*-
import os
import threading
import tkinter as tk # imports
from tkinter import ttk
# from performance_function_test.write_performance_data import getPerformance_data
from tkinter import messagebox
import time
import yaml
import inspect
import ctypes
import xlrd
from PIL import ImageTk
from numpy import *

from openpyxl import load_workbook

# from performance_function_test.write_performance_data import write_performance_data
from performance_function_test.write_performance_data import creat_excel, write_data_to_excel

import sys
sys.path.append("..")
from file_util import *

global projectid,path_report
projectid=1
path_report=get_base_path() + 'performance_function_test\\report'
path_yaml = get_base_path() + 'Folder\performance_function_test_H_folder\yaml'

def init_tab9(tab_page):
    global text_monkey_info
    monty = ttk.LabelFrame(tab_page, text=' ')
    monty.grid(column=0, row=0, padx=20, pady=20)

    # tk.Button(monty, text='点击查看教程 ', font=("Arial", 11), width=30, height=1,command=open_course).grid(column=0, row=0, sticky='W', padx=20, pady=10)


    tk.Button(monty, text='点击配置文件 ', font=("Arial", 11), width=30, height=1,command=lambda: openReport(path_yaml)).grid(column=0, row=1, sticky='W', padx=20, pady=10)
    tk.Button(monty, text='环境配置：点击下载itest APP ', font=("Arial", 11), width=30, height=1,command=install_package).grid(column=0, row=2, sticky='W', padx=20, pady=10)
    tk.Button(monty, text='环境配置：运行adb shell dalvikvm -cp /sdcard/start.dex Start ', font=("Arial", 11), width=45, height=1,command=run_dalvikvm).grid(column=1, row=2, sticky='W', padx=20, pady=10)

    tk.Button(monty, text='点击开始运行 ', font=("Arial", 11), width=30, height=1,command=run_function_test).grid(column=0, row=5, sticky='W', padx=20, pady=10)
    tk.Button(monty, text='查看测试结果 ', font=("Arial", 11), width=30, height=1,command=lambda: openReport(path_report)).grid(column=0, row=9, sticky='W', padx=20, pady=10)

    tk.Label(monty, text="测试过程中遇到的测试场景、欢迎补充：                                              ", fg="blue",font=("Arial", 11), width=50, height=2).grid(column=1, row=12, columnspan=3,sticky='W',
                                                                                   padx=10, pady=2)
    tk.Label(monty, text="1、退出业务场景后，内存下降不明显                                      \n         如：唱唱安卓进入k房选择音效后，退出k房，内存没释放", fg="blue", font=("Arial", 11), width=50, height=2).grid(column=1, row=13, columnspan=3,sticky='W',
                                                                                   padx=10, pady=2)

    tk.Label(monty, text="2、长时间在某一场景下，内存一直缓慢上涨，没有释放内存", fg="blue",font=("Arial", 11), width=50, height=2).grid(column=1, row=14, columnspan=3, sticky='W',padx=10, pady=2)
    tk.Label(monty, text="3、进入某个页面不做任何操作，页面内存飙升又GC再次飙   \n升又GC  如：无效的内存申请，申请了内存没有具体用\n途，然后系统就gc了把无用的都释放了，然后他仍然在\n申请，内存图形呈锯齿状                                                 ", fg="blue",font=("Arial", 11), width=50, height=5).grid(column=1, row=15, columnspan=3, sticky='W',                                                   padx=10, pady=2)
    tk.Label(monty, text="4、新版本比旧版本整体/部分内存上涨超过较多（如20M）    ", fg="blue",font=("Arial", 11), width=50, height=2).grid(column=1, row=16, columnspan=3, sticky='W', padx=10, pady=2)
    # tk.Label(monty, text="5、操作某场景后cpu飙升不下降                                     ", fg="blue", font=("Arial", 11), width=50,height=2).grid(column=1, row=17, columnspan=3, sticky='W',padx=1)
    tk.Label(monty, text="5、操作某场景后cpu飙升不下降                                              ",
             fg="blue", font=("Arial", 11), width=50, height=2).grid(column=1, row=17, columnspan=3, sticky='W',
                                                                     padx=10, pady=2)

    tk.Label(monty, text="6、操作某场景后内存飙升不下降                                            \n如：铃声IOS进入选择照片，照片量很大，内存飙升   ", fg="blue", font=("Arial", 11), width=50,height=2).grid(column=1, row=18, columnspan=3, sticky='W',padx=10, pady=2)


def get_function():
    log_path = get_base_path() + 'performance_function_test\\log\\log.txt'
    path_log = os.path.abspath(log_path)

    path_yaml = get_base_path() + 'Folder\performance_function_test_H_folder\yaml\performance_function_test_H.yaml'
    excel_path = creat_excel(path_yaml)

    with open(path_yaml, 'r', encoding='utf-8') as file:
        data = yaml.load(file)
    scene_info = data['scene']
    process_info = data['process']
    #进程数量
    lens_process = len(process_info)
    #场景数量
    lens_scene = len(scene_info)

    averageValueLog_path = get_base_path() + 'performance_function_test\log\\averageValueLog.txt'

    isCustomStop = False # 是否用户取消
    #初始化所有平均数的列表
    all_average_pss = []
    all_average_cpu = []


    for n in range(1,4):
        len_process = 0
        #初始化，将一组数据的平均值写入到excel中
        average_process = 0
        for temp_scene in scene_info.values():
            # f = open(path_log, 'w')
            # f.closed
            f = open(path_log, 'w')

            start_info = tk.messagebox.askquestion(title='第'+ str(n) +'轮测试：' + temp_scene,message='请对页面进行操作，操作完成后，点击“确定”，进行数据采集')
            print(start_info)
            th = threading.Thread(target=set_text_test, args=(f,))
            if start_info=='yes':
                print('数据采集中')
                th.setDaemon(True)  # 守护线程
                th.start()
            else:
                isCustomStop = True
                break
            end_info=tk.messagebox.showinfo('数据采集中', '正在采集数据，请至少采集30S的数据，\n点击" 确定 "可结束数据采集')

            # tk.Label(monty,text='路径为：', font=("Arial", 11), width=20, height=2).grid(column=1,row=6, sticky='W')
            if end_info =='ok':
               # print('13131233------------------------------------------')
               get_adb_kugou_stop(n,temp_scene,excel_path,len_process,lens_process,average_process,all_average_pss,all_average_cpu)
            len_process = len_process +2 + lens_process + lens_process
            # lens_process = lens_process + 1
            average_process = len_process
        if isCustomStop:
            break

    if not isCustomStop:
        tk.messagebox.showinfo('APP性能测试', '测试已完成，请查看测试结果')

        # file_name = "D:\cx_test_tool\performance_function_test\\report\手机端性能测试结果__2020-02-25-14_08_57.xlsx"

        # 读取excel数据
        file = xlrd.open_workbook(excel_path)

        #写入excel
        workbook = load_workbook(excel_path)
        workbook.guess_types = True  # 猜测格式类型
        sheet_average = workbook.get_sheet_by_name('APP性能测试')

        for j in range(0, (lens_process*2+2)*lens_scene):
            zushu = []
            for test_times in range(1, 4):
                sheet = file.sheet_by_name("第" + str(test_times) + "次测试")

                # 获取第一行的数据
                row_value = sheet.cell_value(2, j+1)
                # print(row_value)
                zushu.append(row_value)
            print('第' + str(j+1) + '组数' + str(zushu))
            pingjun = mean(zushu)
            print(pingjun)
            sheet_average.cell(2 + j // (lens_process*2+2 ), 2 + j % (lens_process*2+2), pingjun)

        workbook.save(excel_path)
        workbook.close()


def set_text_test(f):

    get_adb_info()


def get_adb_kugou_stop(n,temp_scene,excel_path,len_process,lens_process,average_process,all_average_pss,all_average_cpu):
    global get_adb_kugou_isrun
    get_adb_kugou_isrun = False
    stop_get_adb_info(n,temp_scene,excel_path,len_process,lens_process,average_process,all_average_pss,all_average_cpu)

def run_function_test():
    th = threading.Thread(target=get_function())
    th.setDaemon(True)  # 守护线程
    th.start()

#往txt中插入数据
def insert_text(message):
    if(message==None):
        message = ""

    text_monkey_info.insert(tk.END,message)

def openReport(path_report):

    path = os.path.abspath(path_report)  # 获取当前脚本所在的路径
    os.startfile(path)

def _async_raise(tid, exctype):
   """raises the exception, performs cleanup if needed"""
   tid = ctypes.c_long(tid)
   if not inspect.isclass(exctype):
      exctype = type(exctype)
   res = ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, ctypes.py_object(exctype))
   if res == 0:
      raise ValueError("invalid thread id")
   elif res != 1:
      # """if it returns a number greater than one, you're in trouble,
      # and you should call it again with exc=NULL to revert the effect"""
      ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, None)
      raise SystemError("PyThreadState_SetAsyncExc failed")

def get_adb_info():
        path_yaml = get_base_path() + 'Folder\performance_function_test_H_folder\yaml\performance_function_test_H.yaml'
        with open(path_yaml, 'r', encoding='utf-8') as file:
            data = yaml.load(file)
        process_info = data['process']

        process_info_all = ''
        for temp_process_info in process_info.values():
            process_info_all = temp_process_info + ','+ process_info_all
        print(process_info_all[0:-1])

        global get_adb_kugou_isrun
        get_adb_kugou_isrun = True
        #启动itest APP
        # cmd = "adb shell am start -a android.intent.action.MAIN -c android.intent.category.LAUNCHER -n iflytek.testTech.propertytool/.activity.BootActivity"
    #         # os.system(cmd)
        #开启权限
        # time.sleep(2)
        cmd = "adb shell dalvikvm -cp /sdcard/start.dex Start"
        os.system(cmd)
        #开始测试多进程
        cmd = "adb shell am broadcast -a monitorStart --es monitor cpu,pss --es pkg " + process_info_all[0:-1] +"  --ei interval 1000"
        os.system(cmd)

def stop_get_adb_info(n,temp_scene,excel_path,len_process,lens_process,average_process,all_average_pss,all_average_cpu):
        #结束itest采集数据的命令
        cmd = "adb shell am broadcast -a monitorFinish"
        os.system(cmd)

        #获取存放handTest日志路径
        path_handTest_log = get_base_path() + 'Folder\performance_function_test_H_folder\log'
        #将手机上的handtest日志拷贝到电脑上来
        cmd_cpu = "adb pull /storage/emulated/0/AndroidPropertyTool4/handTest " + path_handTest_log
        os.system(cmd_cpu)

        # 获取yaml配置文件路径
        path_yaml_H = get_base_path() + 'Folder\performance_function_test_H_folder\yaml\performance_function_test_H.yaml'
        with open(path_yaml_H,
                  'r', encoding='utf-8') as file:
            data = yaml.load(file)

        process_info = data['process']
        average_pss = []
        average_cpu = []

        workbook = load_workbook(excel_path)
        workbook.guess_types = True  # 猜测格式类型
        sheet = workbook.get_sheet_by_name('第' + str(n) + '次测试')

        for temp_process_info in process_info.values():
            # print(temp_process_info)
            process_list = temp_process_info.split('.')

            path_process = ''
            for temp_process in process_list:
                # print(temp_process)
                path_process = path_process + '_' + temp_process

            # 计算一次测试的数据平均值与总和
            pss_path = "pss" + path_process + ".txt"
            once_pss = []
            print("--------------------------------")

            #获取handTest的txt文件路径
            path_handTest_txt = path_handTest_log + '\handTest\\'


            once_pss_lines = open(path_handTest_txt+ pss_path).readlines()

            #
            # for line in range(2, len(once_pss_lines)):
            #     data = once_pss_lines[line].split()
            #     # print(data[1])
            #     # print(line)
            #     once_pss.append(data[1])

            for line in range(2, len(once_pss_lines)):
                data = once_pss_lines[line].split()
                # print(data[1])
                sheet.cell(2+line, 2 + len_process , data[1])
                once_pss.append(data[1])

            #计算一次测试的内存数据平均值与总和
            print('第'+ str(n) + '轮: ' + '场景'+ temp_scene+ temp_process_info + '进程一组数据的内存值' + str(once_pss))
            sum = 0
            for i in once_pss:
                sum = sum + int(i)
            b = sum / int(len(once_pss))
            print('第'+ str(n) + '轮: ' + '场景'+ temp_scene+ temp_process_info + '进程一组数据的内存平均值是' + str(int(b)))
            average_pss.append(int(b))

            # 计算一次测试的CPU数据平均值与总和
            cpu_path = "cpu" + path_process + ".txt"
            once_cpu = []

            once_cpu_lines = open(
                path_handTest_txt + cpu_path).readlines()

            for temp_once_cpu_lines in range(2, len(once_cpu_lines)):
                data_cpu = once_cpu_lines[temp_once_cpu_lines].split()

                sheet.cell(2 + temp_once_cpu_lines, 2 +1 + lens_process + len_process , data_cpu[1])
                once_cpu.append(data_cpu[1])

            sum = 0
            print('第'+ str(n) + '轮: ' + '场景'+ temp_scene+ temp_process_info + '进程一组数据的CPU值：' + str(once_cpu))
            for i in once_cpu:
                # print(i)
                sum = sum + float(i)

            once_cpu_info = sum / int(len(once_cpu))
            print('第'+ str(n) + '轮: ' + '场景'+ temp_scene+ temp_process_info + '进程一组数据的CPU平均值是' + str(int(once_cpu_info)))
            average_cpu.append(int(once_cpu_info))
            len_process = len_process + 1



        pss_sum_numbers = 0
        for temp_average_pss in average_pss:
            #将一组数据内存平均值写入到excel中
            sheet.cell(3, 2 + average_process , temp_average_pss)
            pss_sum_numbers += temp_average_pss
            average_process = average_process+ 1
        average_pss.append(int(pss_sum_numbers))

        sheet.cell(3, 2 + average_process, pss_sum_numbers)
        average_process = average_process+1
        print('内存平均值列表' + str(average_pss))

        cpu_sum_numbers = 0
        for temp_average_cpu in average_cpu:
            #将一组数据的cpu平均值值写入到excel中
            sheet.cell(3, 2 + average_process, temp_average_cpu)
            cpu_sum_numbers += temp_average_cpu
            average_process = average_process + 1


        average_cpu.append(cpu_sum_numbers)
        # 将一组CPU数据的所有平均值总和写入到excel中
        sheet.cell(3, 2 + average_process, cpu_sum_numbers)
        print('cpu的平均值列表' + str(average_cpu))

        average_process=average_process+1

        all_average_pss.append(average_pss)
        all_average_cpu.append(average_cpu)

        print('所有组数的内存平均值列表：'+ str(all_average_pss))
        print('所有组数的CPU平均值列表：' + str(all_average_cpu))
        workbook.save(excel_path)
        workbook.close()




def install_package_tab():
    file_dir = get_base_path() + 'Folder\performance_function_test_H_folder\package'
    for root, dirs, files in os.walk(file_dir):
        print(files)  # 当前路径下所有非目录子文件
        print(dirs)  # 当前路径下所有子目录
        file_package = file_dir + '\\' + files[0]
        cmd = "adb install " + file_package
        print("运行脚本是：" + cmd)
        os.system(cmd)


def install_package():
    th = threading.Thread(target=install_package_tab)
    th.setDaemon(True)  # 守护线程
    th.start()

def get_dalvikvm():
    cmd = 'adb shell dalvikvm -cp /sdcard/start.dex Start'
    print("运行脚本是：" + cmd)
    assist_result = os.system(cmd)
    print(assist_result)
    if assist_result == 0:
        tk.messagebox.showinfo("Message", 'assist is running')  # 弹出消息窗口
    else:
        tk.messagebox.showinfo("Message", 'assist运行失败，请检查连接')  # 弹出消息窗口


def run_dalvikvm():
    th = threading.Thread(target=get_dalvikvm())
    th.setDaemon(True)  # 守护线程
    th.start()

def open_course_word():
    path = get_base_path() + 'Folder\performance_function_test_H_folder\Course_word\\xn_course.docx'
    open_course = os.system(path)

#开启教程多线程
def open_course():
    th = threading.Thread(target=open_course_word)
    th.setDaemon(True)  # 守护线程
    th.start()



