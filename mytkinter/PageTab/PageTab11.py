# -*- coding: utf-8 -*-
import shutil
import threading
import tkinter as tk # imports
from tkinter import ttk
import os,platform
import subprocess
import re
import time

from openpyxl import load_workbook

# from performance_function_test.changchang.changchang_get_data import mem,cpu,get_adb_data_info
# from performance_function_test.leida.leida_get_data import adb_leida_all

global new_excel_path
new_excel_path = '..\performance_function_test\changchang\酷狗唱唱APP性能测试报告\酷狗唱唱__手机端app性能测试结果__写入模板.xlsx'

def init_tab11(tab_page):
    global text_device_info,text_adb_info
    monty = ttk.LabelFrame(tab_page, text=' ')
    monty.grid(column=0, row=0, padx=20, pady=20)

    tk.Button(monty, text='点击复制测试结果 ', font=("Arial", 11), width=20, height=1, command=creat_leida_excel).grid(column=0, row=0,columnspan=2,
                                                                                                  sticky='W', padx=10)
    tk.Button(monty, text='点击打开实时性能工具 ', font=("Arial", 11), width=20, height=1, command=openTools).grid(
        column=2, row=0, columnspan=2,
        sticky='W', padx=10)

    tk.Label(monty, text="第一次测试", font=("Arial", 11), width=10, height=2).grid(column=0,row=1, sticky='W', padx=20)
    tk.Label(monty, text="    场景1：", font=("Arial", 11), width=10, height=2).grid(column=1,row=1,sticky='W')
    switchscrBtn = tk.Button(monty, text='点击开始 ', fg = 'black',font=("Arial", 11), width=13, height=1 ,command =lambda: run_test(1,1))
    switchscrBtn.grid(column=2, row=1, sticky='W',padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1 , command =lambda: get_adb_kugou_stop(4,2,'第一次测试',1,1)).grid(column=3, row=1,sticky='W',padx=10)

    tk.Label(monty, text="                        场景2：", font=("Arial", 11), width=20, height=2).grid(column=4, row=1, sticky='W')
    switchscrBtn = tk.Button(monty, text='点击开始 ', fg='black',font=("Arial", 11), width=13, height=1,command =lambda: run_test(1,2))
    switchscrBtn.grid(column=5, row=1, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1 , command =lambda: get_adb_kugou_stop(4,12,'第一次测试',1,2)).grid(column=6, row=1, sticky='W', padx=10)

    tk.Label(monty, text="    场景3：", font=("Arial", 11), width=10, height=2).grid(column=1,row=3,sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(1,3)).grid(column=2, row=3, sticky='W',padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,22,'第一次测试',1,3)).grid(column=3, row=3,sticky='W',padx=10)

    tk.Label(monty, text="                        场景4：", font=("Arial", 11), width=20, height=2).grid(column=4, row=3, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(1,4)).grid(column=5, row=3, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,32,'第一次测试',1,4)).grid(column=6, row=3, sticky='W', padx=10)

    tk.Label(monty, text="    场景5：", font=("Arial", 11), width=10, height=2).grid(column=1, row=4, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(1,5)).grid(column=2, row=4, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,42,'第一次测试',1,5)).grid(column=3, row=4, sticky='W', padx=10)


    tk.Label(monty, text="                        场景6：", font=("Arial", 11), width=20, height=2).grid(column=4, row=4,sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1, command=lambda: run_test(2, 6)).grid(column=5, row=4, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command=lambda: get_adb_kugou_stop(4, 52, '第一次测试', 2, 6)).grid(column=6, row=4, sticky='W', padx=10)
    tk.Label(monty, text="          ", font=("Arial", 11), width=20, height=2).grid(column=4, row=5, sticky='W')



    tk.Label(monty, text="第二次测试", font=("Arial", 11), width=10, height=2).grid(column=0, row=6, sticky='W', padx=20)
    tk.Label(monty, text="    场景1：", font=("Arial", 11), width=10, height=2).grid(column=1, row=6, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(2,1)).grid(column=2, row=6, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,2,'第二次测试',2,1)).grid(column=3, row=6, sticky='W', padx=10)

    tk.Label(monty, text="                        场景2：", font=("Arial", 11), width=20, height=2).grid(column=4, row=6,
                                                                                                      sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(2,2)).grid(column=5, row=6, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,12,'第二次测试',2,2)).grid(column=6, row=6, sticky='W', padx=10)

    tk.Label(monty, text="    场景3：", font=("Arial", 11), width=10, height=2).grid(column=1, row=7, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(2,3)).grid(column=2, row=7, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,22,'第二次测试',2,3)).grid(column=3, row=7, sticky='W', padx=10)

    tk.Label(monty, text="                        场景4：", font=("Arial", 11), width=20, height=2).grid(column=4, row=7,
                                                                                                      sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(2,4)).grid(column=5, row=7, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,32,'第二次测试',2,4)).grid(column=6, row=7, sticky='W', padx=10)

    tk.Label(monty, text="    场景5：", font=("Arial", 11), width=10, height=2).grid(column=1, row=8, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(2,5)).grid(column=2, row=8, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,42,'第二次测试',2,5)).grid(column=3, row=8, sticky='W', padx=10)

    tk.Label(monty, text="                        场景6：", font=("Arial", 11), width=20, height=2).grid(column=4, row=8,  sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1, command=lambda: run_test(2, 6)).grid(column=5, row=8, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1, command=lambda: get_adb_kugou_stop(4, 52, '第二次测试', 2, 6)).grid(column=6, row=8, sticky='W', padx=10)

    tk.Label(monty, text="          ", font=("Arial", 11), width=20, height=2).grid(column=4, row=9,
                                                                                    sticky='W')



    tk.Label(monty, text="第三次测试", font=("Arial", 11), width=10, height=2).grid(column=0, row=11, sticky='W', padx=20)
    tk.Label(monty, text="    场景1：", font=("Arial", 11), width=10, height=2).grid(column=1, row=11, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(3,1)).grid(column=2, row=11, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,2,'第三次测试',3,1)).grid(column=3, row=11, sticky='W', padx=10)

    tk.Label(monty, text="                        场景2：", font=("Arial", 11), width=20, height=2).grid(column=4, row=11,
                                                                                                      sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(3,2)).grid(column=5, row=11, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,12,'第三次测试',3,2)).grid(column=6, row=11, sticky='W', padx=10)

    tk.Label(monty, text="    场景3：", font=("Arial", 11), width=10, height=2).grid(column=1, row=12, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(3,3)).grid(column=2, row=12, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,22,'第三次测试',3,3)).grid(column=3, row=12, sticky='W', padx=10)

    tk.Label(monty, text="                        场景4：", font=("Arial", 11), width=20, height=2).grid(column=4, row=12,
                                                                                                      sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(3,4)).grid(column=5, row=12, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,32,'第三次测试',3,4)).grid(column=6, row=12, sticky='W', padx=10)

    tk.Label(monty, text="    场景5：", font=("Arial", 11), width=10, height=2).grid(column=1, row=13, sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1,command =lambda: run_test(3,5)).grid(column=2, row=13, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command =lambda: get_adb_kugou_stop(4,42,'第三次测试',3,5)).grid(column=3, row=13, sticky='W', padx=10)



    tk.Label(monty, text="                        场景6：", font=("Arial", 11), width=20, height=2).grid(column=4, row=13,sticky='W')
    tk.Button(monty, text='点击开始 ', font=("Arial", 11), width=13, height=1, command=lambda: run_test(2, 6)).grid(column=5, row=13, sticky='W', padx=10)
    tk.Button(monty, text='点击结束 ', font=("Arial", 11), width=13, height=1,command=lambda: get_adb_kugou_stop(4,52, '第三次测试', 2, 6)).grid(column=6, row=13, sticky='W', padx=10)
    tk.Label(monty, text="          ", font=("Arial", 11), width=20, height=2).grid(column=4, row=14, sticky='W')

    text_adb_info = tk.Text(monty, width=70, height=5, font=("Arial", 12))
    text_adb_info.grid(row=16, column=0, sticky='W', columnspan=7, padx=20, pady=10)

    tk.Button(monty, text='点击查看测试结果 ', font=("Arial", 11), width=20, height=1 ,command=openReport).grid(column=5, row=16, sticky='W',columnspan=2, padx=10,pady=15)

    tk.Label(monty, text="          ", font=("Arial", 11), width=20, height=2).grid(column=4, row=17,
                                                                                    sticky='W')

def openReport():
    path_report = '..\performance_function_test\leida\浮浮雷达APP性能测试报告'
    path = os.path.abspath(path_report)  # 获取当前脚本所在的路径
    os.startfile(path)

def openTools():
    path_report = '..\Tools\performance_tools'
    path = os.path.abspath(path_report)  # 获取当前脚本所在的路径
    os.startfile(path)

def set_text_test():
    global get_adb_kugou_isrun
    get_adb_kugou_isrun = True

    new_path = '..\mytkinter\log\log_new.txt'
    path_log = os.path.abspath(new_path)
    f = open(path_log, 'w')

    while get_adb_kugou_isrun:
        adb_leida_all(f)
    f.close()

def run_test(n,y):
    th = threading.Thread(target=set_text_test)
    th.setDaemon(True)  # 守护线程
    th.start()
    insert_text(str(get_time()) + '第 '+ str(n) + ' 次测试__'+'场景 '+str(y)+' 测试开始了\n')


# def run_test(switchscrBtn,n,y):
#     if  switchscrBtn['text'] == '点击开始':
#         switchscrBtn['text'] = '测试开始了'
#         switchscrBtn['fg']= ('blue')
#
#         th = threading.Thread(target=set_text_test)
#         th.setDaemon(True)  # 守护线程
#         th.start()
#         insert_text(str(get_time()) + '第 '+ str(n) + ' 次测试__'+'场景 '+str(y)+' 测试开始了\n')
#
#     else:
#         switchscrBtn['text'] = '点击开始'


def get_adb_kugou_stop(begin_row,begin_col,sheet_name,n,y):
    global get_adb_kugou_isrun
    get_adb_kugou_isrun = False
    time.sleep(10)
    write_leida_excel(new_excel_path,begin_row,begin_col,sheet_name)
    insert_text(str(get_time()) + '第 ' + str(n) + ' 次测试__' + '场景 ' + str(y) + ' 测试结束了\n')


def creat_leida_excel():
    import time
    global new_excel_path
    model_excel_path = os.path.abspath('..\performance_function_test\leida\浮浮雷达APP性能测试报告\浮浮雷达__手机端app性能测试结果__模板.xlsx')

    time_report = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
    excel_result = shutil.copy(model_excel_path, u'..\performance_function_test\leida\浮浮雷达APP性能测试报告\浮浮雷达__手机端app性能测试报告__' + str(time_report) + '.xlsx')


    new_excel_path=os.path.abspath(excel_result)
    # 删除文件
    if os.path.exists(new_excel_path):
        os.unlink(new_excel_path)
    # 复制并重命名新文件
    shutil.copy(model_excel_path, new_excel_path)
    # time_log = time.strftime(" %Y_%m_%d_%H_%M_%S:   ", time.localtime(time.time()))
    insert_text(str(get_time())+'创建测试结果excele成功\n')
    # return new_excel_path

def get_time():
    time_log = time.strftime(" %Y_%m_%d_%H_%M_%S:   ", time.localtime(time.time()))
    return time_log

def write_data_to_excel(new_txt_path,excel_path,begin_row, begin_col,sheet_name):
    '''
    :param old_txt_path:  新的的txt数据值路径
    :param excel_path: excel文件绝对路径
    :param begin_row: 写入数据开始行号，最小为1
    :param begin_col: 写入数据开始列数，最小为1
    :return:
    '''
    file_txt = open(new_txt_path)
    lines = file_txt.readlines()

    if (begin_row <= 1):
        begin_row = 1
    if (begin_col <= 1):
        begin_col = 1

    workbook = load_workbook(excel_path)
    workbook.guess_types = True  # 猜测格式类型
    sheet = workbook.get_sheet_by_name(sheet_name)


    if (len(lines) == 0):
        print("没有数据.......")
        return


    for i in range(len(lines)):
        list = (lines[i].split(','))
        if (len(list) == 0):
            continue

        for j in range(len(list)):
            # print('temp的数据 ' + list[j])
            sheet.cell(i + begin_row, j + begin_col, float(list[j]))
    workbook.save(excel_path)
    workbook.close()
    file_txt.close()
    return lines,sheet

def write_leida_excel(new_excel_path,begin_row,begin_col,sheet_name):
    new_txt_path = os.path.abspath('..\mytkinter\log\log_new.txt')

    excel_path = new_excel_path
    write_data_to_excel(new_txt_path, excel_path, begin_row, begin_col,sheet_name)


    print('写入成功')


#往txt中插入数据
def insert_text(message):
    if(message==None):
        message = ""
    text_adb_info.insert(1.0,message)

if __name__ == "__main__":
    get_data_info()